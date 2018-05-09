import csv
from datetime import datetime
import itertools
import os
import re
from tempfile import NamedTemporaryFile
from xml.etree import ElementTree as ET

from lxml import etree
import unicodedata
from openpyxl import load_workbook
from pyxform.utils import has_external_choices
from mangrove.datastore.entity_type import entity_type_already_defined
from mangrove.datastore.queries import get_non_voided_entity_count_for_type
from pyxform.xls2json import parse_file_to_json
import xlrd
import xmldict
from pyxform import create_survey_element_from_dict
from datawinners.activitylog.models import UserActivityLog
from datawinners.common.constant import CREATED_QUESTIONNAIRE

from datawinners.project.wizard_view import create_questionnaire
from datawinners.accountmanagement.models import NGOUserProfile
from datawinners.main.database import get_database_manager
from datawinners.project.helper import generate_questionnaire_code, associate_account_users_to_project
from mangrove.datastore.user_permission import grant_user_permission_for
from mangrove.form_model.field import FieldSet, GeoCodeField, DateField, PhotoField, VideoField, AudioField, MediaField, \
    SelectField, DateTimeField


# used for edit questionnaire in xls questionnaire flow
# noinspection PyUnresolvedReferences
from datawinners.search import *
from mangrove.transport.player.parser import XlsParser
from django.utils.translation import ugettext as _
from xml.sax.saxutils import escape


def get_generated_xform_id_name(xform):
    xform_cleaned = re.sub(r"\s+", " ", re.sub(r"\n", "", xform))
    match = re.search('<model>.* <instance> <(.+?) id="', xform_cleaned)
    return match.group(1)


CALCULATE = 'calculate'
BARCODE = 'barcode'
DEVICEID = 'deviceid'
SUBSCRIBERID = 'subscriberid'
IMEI_ID = 'imei'
PHONE_NUMBER = 'phonenumber'
TODAY = 'today'
START = 'start'
END = 'end'
SIMSERIAL = 'simserial'


def encode_xls_value(value, value_type):
    if value_type == float:
        int_value = int(value)
        if int_value == value:
            return unicode(int_value)
        else:
            return unicode(value)
    else:
        return value.encode('utf-8')


class XlsFormParser():
    type_dict = {'group': ['repeat', 'group'],
                 'field': ['text', 'integer', 'decimal', 'date', 'geopoint', 'calculate', 'cascading_select', BARCODE,
                           'time', 'datetime', 'dw_idnr', DEVICEID, SUBSCRIBERID, IMEI_ID, PHONE_NUMBER, TODAY, START, END, SIMSERIAL],
                 'auto_filled': ['note'],
                 'media': ['photo', 'audio', 'video'],
                 'select': ['select one', 'select all that apply', 'select one or specify other',
                            'select all that apply or specify other'],
                 'select_external' : ['select one external']
    }
    recognised_types = list(itertools.chain(*type_dict.values()))
    supported_types = [type for type in recognised_types if type not in type_dict['auto_filled']]
    or_other_data_types = ['select all that apply or specify other', 'select one or specify other']
    select_without_list_name = ['select_one', 'select_multiple']

    def __init__(self, path_or_file, questionnaire_name, dbm=None, excel_raw_stream=None, file_type=None):
        self.questionnaire_name = questionnaire_name
        self.dbm = dbm
        if excel_raw_stream is not None:
          self._file_object = excel_raw_stream
          self.path = 'questionnaire.'+file_type #Used only to deduct the extension internally in pyxform   
        elif isinstance(path_or_file, basestring):
            self._file_object = None
            self.path = path_or_file
        else:
            self._file_object = path_or_file
            self.path = path_or_file.name

        self.xform_dict = parse_file_to_json(self.path, file_object=self._file_object)

    def _validate_for_uppercase_names(self, field):
        if filter(lambda x: x.isupper(), field['name']):
            return _("Uppercase in names not supported")
        return None

    def _is_multi_language(self):
        for child in self.xform_dict['children']:
            if isinstance(child.get('label'), dict) and len(child.get('label', {})) > 1:
                return True
        return False
        # return len(self.xform_dict['children'][0]['label']) > 1

    def _identify_default_language(self):
        if self.xform_dict['default_language'] != u'default':
            self.default_language = self.xform_dict['default_language']
            return

        if not self._has_explicit_language_specified(self.xform_dict['children']):
            # avoid loading excel again if not multi language questionnaire
            self.default_language = 'default'
            return

        if self.path.endswith('.xls'):
            self._identify_language_from_xls_file()
        else:
            self._identify_language_from_xlsx_file()

    def _identify_language_from_xls_file(self):
        headers = xlrd.open_workbook(filename=self.path).sheet_by_name('survey').row(0)
        for header_cell in headers:
            if re.match('^label::', header_cell.value):
                language = header_cell.value.split("::")[1]
                self.default_language = language
                return

    def _identify_language_from_xlsx_file(self):
        try:
            workbook = load_workbook(self.path)
        except Exception as e:
            raise e
        worksheet = workbook.get_sheet_by_name('survey')

        for counter, row in enumerate(worksheet.iter_rows()):
            if counter > 0:
                break

            for cell in row:
                if cell.value and re.match('^label::', cell.value):
                    language = cell.value.split("::")[1]
                    self.default_language = language
                    return

    def _create_question(self, field, parent_field_code=None):
        unique_id_errors = []
        question = None
        errors = []
        field_type = field['type'].lower()
        if field_type in self.type_dict['group']:
            question, errors, unique_id_errors = self._group(field, parent_field_code)
        elif field_type in self.type_dict['field']:
            question = self._field(field, parent_field_code)
        elif field_type in self.type_dict['select']:
            question = self._select(field, parent_field_code)
        elif field_type in self.type_dict['select_external']:
            question = self._select_external(field, parent_field_code)
        elif field_type in self.type_dict['media']:
            question = self._media(field, parent_field_code)
        return question, errors, unique_id_errors

    def _create_questions(self, fields, parent_field_code=None):
        questions = []
        errors = []
        info = []
        for field in fields:
            if field.get('control', None) and field['control'].get('bodyless', False):  # ignore calculate type
                continue
            if field['type'].lower() in self.supported_types:
                try:
                    question, field_errors, field_info = self._create_question(field, parent_field_code)

                    if not field_errors and question:
                        questions.append(question)

                        if question['type'] in ['select', 'select1'] and question['has_other']:
                            other_field = {u'type': u'text', u'name': question['code'] + "_other",
                                           u'label': question['title'] + "_other", u'is_other': True}
                            other_question, other_field_errors, x = self._create_question(other_field,
                                                                                          parent_field_code)
                            questions.append(other_question)

                        if question['type'] == 'unique_id':
                            self.validate_unique_id_type(field)

                    else:
                        errors.extend(field_errors)
                    info.extend(field_info)
                except LabelForChoiceNotPresentException as e:
                    errors.append(e.message)
                except UniqueIdNotFoundException as e:
                    errors.append(e.message)
                except UniqueIdNotMentionedException as e:
                    errors.append(e.message)
                except UniqueIdNumbersNotFoundException as e:
                    info.append(e.message)
                except LabelForFieldNotPresentException as e:
                    errors.append(e.message)
                except ForbiddenWordInFieldNameException as e:
                    errors.append(e.message)
                except ErrorSyntaxInRelevantException as e:
                    errors.append(e.message)
        return questions, set(errors), set(info)

    def _validate_group(self, errors, field):
        if field['type'] == 'repeat':
            try:
                self._validate_for_nested_repeats(field)
            except NestedRepeatsNotSupportedException as e:
                errors.append(e.message)
        child_errors = self._validate_fields_are_recognised(field['children'])
        errors.extend(child_errors)

    def _validate_fields_are_recognised(self, fields):
        errors = []
        for field in fields:
            # try:
            #     self._validate_for_no_language(field)
            # except MultipleLanguagesNotSupportedException as e:
            #     errors.append(e.message)
            field_type = field['type'].lower()
            if field_type in self.recognised_types:
                if field_type in self.type_dict['group']:
                    self._validate_group(errors, field)
                errors.append(self._validate_for_prefetch_csv(field))
            else:
                if (field["type"]) in self.select_without_list_name:
                    errors.append(_("missing list reference, check your select_one or select multiple question types"))
                else:
                    errors.append(_("%s as a datatype") % _(field_type))
            if field.get('media'):
                for media_type in field['media'].keys():
                    errors.append(_("attaching media to fields is not supported %s") % media_type)
        return set(errors) - set([None])

    def _validate_for_nested_repeats(self, field):
        for f in field["children"]:
            if f["type"] == "repeat":
                raise NestedRepeatsNotSupportedException()
            if f["type"] == "group":
                self._validate_for_nested_repeats(f)

    def _has_explicit_language_specified(self, fields):
        for field in fields:
            if self._has_languages(field.get('label')):
                return True
        return False
        #    if self._has_languages(field.get(header)):
        #        raise MultipleLanguagesNotSupportedException()
        # field.get("choices") and self._validate_for_no_language(field.get("choices")[0])
        # if field.get('bind') and self._has_languages(field.get('bind').get('jr:constraintMsg')):
        #     raise MultipleLanguagesNotSupportedException()

    def _has_languages(self, header):
        return header and isinstance(header, dict) and len(header) >= 1

    @staticmethod
    def _get_media_in_choices(choices):
        for choice in choices:
            if choice.get('media'):
                return choice['media'].keys()
        return []

    def _validate_media_in_choices(self, fields):
        errors = []
        for field in fields:
            if field['type'] in self.type_dict['group']:
                choice_errors = self._validate_media_in_choices(field['children'])
                [errors.append(choice_error) for choice_error in choice_errors if choice_error]
            choices = field.get('choices')
            if choices:
                media_type_in_choices = self._get_media_in_choices(choices)
                for media_type in media_type_in_choices:
                    errors.append(_("attaching media to choice fields not supported %s") % _(media_type))
        return errors

    def _validate_default_value(self, errors, field, name_set):
        default_value = field.get('default')
        if default_value:
            default_values_list = default_value.split(' ')
            default_values_set = set(default_values_list)
            if not default_values_set.issubset(name_set):
                errors.append(_(
                    "Entered default value is not defined in the choices list."))

    def _validate_choice_names(self, fields):
        errors = []
        for field in fields:
            if field['type'] in self.type_dict['group']:
                errors.extend(self._validate_choice_names(field['children']))
            choices = field.get('choices')
            if choices:
                # name_list = [choice['name'].lower() for choice in choices]
                name_list = [choice['name'] for choice in choices]
                name_set = set(name_list)
                name_list_without_duplicates = list(name_set)
                self._validate_default_value(errors, field, name_set)
                if len(name_list) != len(name_list_without_duplicates):
                    errors.append(_("duplicate names within list [%s] (choices sheet)") % choice['name'])
                if filter(lambda name: " " in unicode(name), name_list):
                    errors.append(_("spaces in name column (choice sheet)"))
        return errors

    def parse(self):
        fields = self.xform_dict['children']
        errors = self._validate_fields_are_recognised(fields)
        settings_page_errors = self._validate_settings_page_is_not_present()
        errors = errors.union(settings_page_errors)
        choice_errors = self._validate_media_in_choices(fields)
        [errors.add(choice_error) for choice_error in choice_errors if choice_error]
        choice_name_errors = self._validate_choice_names(fields)
        errors = errors.union(set(choice_name_errors))
        self._identify_default_language()
        questions, question_errors, info = self._create_questions(fields)
        if question_errors:
            errors = errors.union(question_errors)
        if not errors and not questions:
            errors.add("Uploaded file is empty!")
        if errors:
            return XlsParserResponse(errors)
        _map_unique_id_question_to_select_one(self.xform_dict)
        survey = create_survey_element_from_dict(self.xform_dict)
        itemsets_csv = None
        if has_external_choices(self.xform_dict):
            itemsets_csv = NamedTemporaryFile(delete=False, suffix=".csv")
            choices_exported = self.sheet_to_csv(self.path, itemsets_csv.name, "external_choices")
            if not choices_exported:
                errors.add("Could not export itemsets.csv, perhaps the external choices sheet is missing.")
            else:
                print 'External choices csv is located at:', itemsets_csv
        xform = survey.to_xml()
        # encoding is added to support ie8
        xform = re.sub(r'<\?xml version="1.0"\?>', '<?xml version="1.0" encoding="utf-8"?>', xform)
        updated_xform = self.update_xform_with_questionnaire_name(xform)
        return XlsParserResponse([], updated_xform, questions, info, self._is_multi_language(), itemsets_csv)

    def sheet_to_csv(self, workbook_path, csv_path, sheet_name):
        wb = xlrd.open_workbook(workbook_path)
        try:
            sheet = wb.sheet_by_name(sheet_name)
        except xlrd.biffh.XLRDError:
            return False
        if not sheet or sheet.nrows < 2:
            return False
        with open(csv_path, 'wb') as f:
            writer = csv.writer(f, quoting=csv.QUOTE_ALL)
            mask = [v and len(v.strip()) > 0 for v in sheet.row_values(0)]
            for r in range(sheet.nrows):
                writer.writerow(
                    [encode_xls_value(v, type(v)) for v, m in zip(sheet.row_values(r), mask) if m])
        return True


    def update_xform_with_questionnaire_name(self, xform):
        # Escape <, > and & and convert accented characters to equivalent non-accented characters
        return re.sub(r"<h:title>\w+</h:",
                      "<h:title>%s</h:" % unicodedata.normalize('NFD', escape(unicode(self.questionnaire_name))).encode(
                          'ascii', 'ignore'), xform)

    def _get_label(self, field):

        if 'label' not in field:
            if field['type'] == 'group' and 'control' in field:
                if field['control']['appearance'] == 'field-list':
                    return field['name']
            elif field['type'] in ['calculate', DEVICEID, SUBSCRIBERID, IMEI_ID, PHONE_NUMBER, TODAY, START, END, SIMSERIAL]:
                return field['name']
            else:
                raise LabelForFieldNotPresentException(field_name=field['name'])

        if field['name'].lower() in ['form']:
            raise ForbiddenWordInFieldNameException(field_name=field['name'])

        if isinstance(field['label'], dict):
            if field['label'].get(self.default_language):
                return field['label'].get(self.default_language)
            raise LabelForFieldNotPresentException(field_name=field['name'])
        else:
            return field['label']

    def _get_relevant(self, field):
        relevant = field.get("bind").get("relevant") if field.get("bind") else None
        if relevant:
            vanilla_pattern = r"^\$\{\w+\}(>|>=|=|<=|<)(\d+|'\w+')$"
            selected_pattern = r"^selected\(\$\{\w+\},( |)(\d+|'\w+'|'\d+')\)$"
            if 'and' in relevant or 'or' in relevant:
                statements = re.split(' and | or ', relevant)
                for phrase in statements:
                    if 'selected' in phrase:
                        pattern = selected_pattern
                    else:
                        pattern = vanilla_pattern
                    if not re.match(pattern, phrase):
                        raise ErrorSyntaxInRelevantException(field_name=relevant, specific=phrase)
            else:
                if 'selected' in relevant:
                    pattern = selected_pattern
                else:
                    pattern = vanilla_pattern
                if not re.match(pattern, relevant):
                    raise ErrorSyntaxInRelevantException(field_name=relevant, specific=None)
        return relevant

    def _group(self, field, parent_field_code=None):
        group_label = self._get_label(field)

        fieldset_type = 'entity'

        if field['type'] == 'repeat':
            fieldset_type = 'repeat'
        elif field['type'] == 'group':
            fieldset_type = 'group'

        name = field['name']
        questions, errors, unique_id_errors = self._create_questions(field['children'], field['name'])
        question = {
            'title': group_label,
            'type': 'field_set',
            "is_entity_question": False,
            "code": name, "name": group_label,
            "required": False,
            "parent_field_code": parent_field_code,
            "instruction": "No answer required",
            "fieldset_type": fieldset_type,
            "fields": questions,
            "appearance": self._get_appearance(field),
            "default": field.get('default'),
            "relevant": field.get("bind").get("relevant") if field.get("bind") else None
        }
        return question, errors, unique_id_errors

    def _get_date_format(self, field):

        if field['type'] == 'time':
            return "HH:mm"

        appearance = self._get_appearance(field)
        if appearance:
            if 'month-year' in appearance:
                return 'mm.yyyy'
            if 'year' in appearance:
                return 'yyyy'


        if field['name'] == 'today':
            return 'dd.mm.yyyy'

        return 'dd.mm.yyyy'

    def _get_unique_id_type(self, field):
        try:
            return field['bind']['constraint']
        except KeyError:
            raise UniqueIdNotMentionedException(field.get('name', ''))

    def validate_unique_id_type(self, field):
        unique_id_type = self._get_unique_id_type(field)
        if not entity_type_already_defined(self.dbm, [unique_id_type]):
            raise UniqueIdNotFoundException(unique_id_type)
        if not get_non_voided_entity_count_for_type(self.dbm, unique_id_type.lower()):
            raise UniqueIdNumbersNotFoundException(unique_id_type)

    def _field(self, field, parent_field_code=None):
        xform_dw_type_dict = {'geopoint': 'geocode', 'decimal': 'integer', CALCULATE: 'text', BARCODE: 'text',
                              'dw_idnr': 'unique_id', DEVICEID: 'text', SUBSCRIBERID: 'text', IMEI_ID: 'text',
                              PHONE_NUMBER: 'text', SIMSERIAL: 'text', START: 'datetime', END: 'datetime', TODAY: 'date'}
        help_dict = {'text': 'word', 'integer': 'number', 'decimal': 'decimal or number', CALCULATE: 'calculated field',
                     'dw_idnr': 'Identification Number'}
        name = self._get_label(field)
        code = field['name']
        type = field['type'].lower()
        hint = field.get("hint")
        constraint_message = field.get("bind").get("jr:constraintMsg") if field.get("bind") else None
        appearance = self._get_appearance(field)
        default = field.get('default')
        xform_constraint = field.get("bind").get("constraint") if field.get("bind") else None
        relevant = self._get_relevant(field)
        question = {'title': name, 'type': xform_dw_type_dict.get(type, type), "is_entity_question": False,
                    "code": code, "name": name, 'required': self.is_required(field), "hint": hint,
                    "constraint_message": constraint_message, "parent_field_code": parent_field_code, "appearance": appearance, "default": default,
                    "xform_constraint": xform_constraint, "relevant": relevant,
                    "instruction": "Answer must be a %s" % help_dict.get(type, type)}  # todo help text need improvement

        if type in ['date', TODAY]:
            format = self._get_date_format(field)
            question.update({'date_format': format, 'event_time_field_flag': False,
                             "instruction": "Answer must be a date in the following format: day.month.year. Example: 25.12.2011"})

        if type == 'dw_idnr':
            question.update({'uniqueIdType': self._get_unique_id_type(field), "is_entity_question": True})

        if type == CALCULATE:
            question.update({"is_calculated": True})

        if field.get('is_other') is not None and field.get('is_other'):
            question.update({"is_other": True})

        return question

    def _get_appearance(self, field):
        if field.get('control') and field['control'].get('appearance'):
            return field['control']['appearance']

    def _select(self, field, parent_field_code=None):
        if self._get_appearance(field) == 'label':
            return
        name = self._get_label(field)
        code = field['name']
        hint = field.get("hint")
        constraint_message = field.get("bind").get("jr:constraintMsg") if field.get("bind") else None
        appearance = self._get_appearance(field)
        default = field.get('default')
        relevant = field.get("bind").get("relevant") if field.get("bind") else None
        xform_constraint = field.get("bind").get("constraint") if field.get("bind") else None

        if field.get('choices'):
            choices = [{'value': {'text': self._get_choice_label(f), 'val': f['name']}} for f in field.get('choices')]
        else:
            # cascade select
            choices = [{'value': {'text': self._get_choice_label(f), 'val': f['name']}} for f in
                       self.xform_dict['choices'].get(field['itemset'])]
        question = {"title": name, "code": code, "type": "select", 'required': self.is_required(field),
                    "hint": hint, "constraint_message": constraint_message, "parent_field_code": parent_field_code,
                    "appearance": appearance, "default": default, "choices": choices, "is_entity_question": False,
                    "xform_constraint": xform_constraint, "relevant": relevant, "is_cascade": field.get("itemset") is not None}

        question.update({"has_other": field['type'] in self.or_other_data_types})

        if field['type'] in ['select one', 'select one or specify other']:
            question.update({"type": "select1"})

        return question

    def _select_external(self, field, parent_field_code=None):
        if self._get_appearance(field) == 'label':
            return
        name = self._get_label(field)
        code = field['name']

        query = field.get('query')
        choice_filter = field.get('choice_filter')
        question = {"title": name, "code": code, "type": "select one external", 'required': self.is_required(field),
                    "parent_field_code": parent_field_code,
                    "query": query, "choice_filter":choice_filter, "is_entity_question": False}
        return question

    def _get_choice_label(self, choice_field):
        if not choice_field.get('label', None):
            raise LabelForChoiceNotPresentException(choice_field.get('name', ''))

        choice_label = choice_field.get('label')
        if not isinstance(choice_label, dict):
            return choice_label

        choice_label = choice_label.get(self.default_language)
        if choice_label:
            return choice_label
        else:
            raise LabelForChoiceNotPresentException(choice_field.get('name', ''))

    def is_required(self, field):
        if field.get('bind') and 'yes' == str(field['bind'].get('required')).lower():
            return True
        return False

    def _media(self, field, parent_field_code=None):
        name = self._get_label(field)
        code = field['name']
        hint = field.get("hint")
        constraint_message = field.get("bind").get("jr:constraintMsg") if field.get("bind") else None
        appearance = self._get_appearance(field)
        default = field.get('default')
        xform_constraint = field.get("bind").get("constraint") if field.get("bind") else None
        relevant = field.get("bind").get("relevant") if field.get("bind") else None
        question = {"title": name, "code": code, "type": field['type'], 'required': self.is_required(field),
                    "parent_field_code": parent_field_code, "hint": hint, "constraint_message": constraint_message,
                    "is_entity_question": False, "appearance": appearance, "default": default, "xform_constraint": xform_constraint, "relevant": relevant}
        return question

    def _validate_for_prefetch_csv(self, field):
        if 'bind' in field and 'calculate' in field['bind'] and 'pulldata(' in field['bind']['calculate']:
            return _("Prefetch of csv not supported")
        return None

    def _validate_settings_page_is_not_present(self):
        errors = []
        not_allowed_settings = ['public_key', 'submission_url', 'instance_name', 'style']
        setting_page_error = _(
            "columns other than Default_Language in the Settings Sheet.")
        if self.xform_dict['title'] != self.xform_dict['name']:
            errors.append(setting_page_error)

        if self.xform_dict['id_string'] != self.xform_dict['name']:
            errors.append(setting_page_error)

        for setting in not_allowed_settings:
            if setting in self.xform_dict:
                errors.append(setting_page_error)
                break
        return set(errors)


def _map_unique_id_question_to_select_one(xform_dict):
    for field in xform_dict['children']:
        if field['type'] == "dw_idnr":
            field['type'] = u'select one'
            field[u'choices'] = [{u'name': field['name'], u'label': u'placeholder'}]
            del field['bind']['constraint']
        elif field['type'] in ['group', 'repeat']:
            _map_unique_id_question_to_select_one(field)


class MangroveService():
    def __init__(self, request, questionnaire_code=None, project_name=None, xls_form=None, xls_parser_response=None):
        self.request = request
        self.user = request.user
        user_profile = NGOUserProfile.objects.get(user=self.user)
        self.reporter_id = user_profile.reporter_id
        self.manager = get_database_manager(self.user)
        self.entity_type = ['reporter']
        self.questionnaire_code = questionnaire_code if questionnaire_code else generate_questionnaire_code(
            self.manager)
        self.name = 'Xlsform-Project-' + self.questionnaire_code if not project_name else project_name
        self.language = 'en'
        self.xform = xls_parser_response.xform_as_string
        self.xform_with_form_code = self.add_form_code(self.questionnaire_code)
        self.json_xform_data = xls_parser_response.json_xform_data
        self.xls_form = xls_form
        self.itemsets_csv = xls_parser_response.csv_path

    def _add_model_sub_element(self, root, name, value):
        generated_id = get_generated_xform_id_name(self.xform)
        model = '{http://www.w3.org/2002/xforms}%s' % generated_id
        [self._add(r, name, value) for r in root.iter(model)]

        node_set = '/%s/%s' % (generated_id, name)
        [ET.SubElement(r, '{http://www.w3.org/2002/xforms}bind', {'nodeset': node_set, 'type': "string"}) for r in
         root.getiterator() if
         r.tag == '{http://www.w3.org/2002/xforms}model']

    def _add(self, instance, name, value):
        e = ET.SubElement(instance, '{http://www.w3.org/2002/xforms}%s' % name)
        e.text = value
        return e

    def add_form_code(self, form_code):
        ET.register_namespace('', 'http://www.w3.org/2002/xforms')
        root = ET.fromstring(self.xform.encode('utf-8'))
        self._add_model_sub_element(root, 'form_code', form_code)
        return '<?xml version="1.0"?>%s' % ET.tostring(root)

    def create_project(self):
        project_json = {'questionnaire-code': self.questionnaire_code}

        questionnaire = create_questionnaire(post=project_json, manager=self.manager, name=self.name,
                                             language=self.language,
                                             reporter_id=self.reporter_id, question_set_json=self.json_xform_data,
                                             xform=self.xform_with_form_code)

        if not questionnaire.is_project_name_unique():
            return None, None

        associate_account_users_to_project(self.manager, questionnaire)
        questionnaire.update_media_field_flag()
        questionnaire.update_doc_and_save()
        grant_user_permission_for(user_id=self.request.user.id, questionnaire_id=questionnaire.id, manager=self.manager)
        if self.xls_form:
            base_name, extension = os.path.splitext(self.xls_form.name)
        else:
            extension = ".xls"
        questionnaire.add_attachments(self.xls_form, 'questionnaire%s' % extension)
        if self.itemsets_csv:
            questionnaire.add_attachments(self.itemsets_csv, "itemsets.csv")
        UserActivityLog().log(self.request, action=CREATED_QUESTIONNAIRE, project=questionnaire.name,
                              detail=questionnaire.name)
        return questionnaire.id, questionnaire.form_code


def _is_choice_item_in_choice_list(item, options):
    for choice_item in options:
        if choice_item['val'] == item:
            return True
    return False


class XFormSubmissionProcessor():
    def get_dict(self, field, value):
        if not value:
            return {field.code: ''}
        if type(field) is FieldSet:
            dicts = []
            for v in value:
                dict = {}
                for f in field.fields:
                    dict.update(self.get_dict(f, v.get(f.code, '')))
                dicts.append(dict)
            return {field.code: dicts}
        elif type(field) is DateField:
            return {field.code: '-'.join(value.split('.')[::-1])}
        elif type(field) is GeoCodeField:
            return {field.code: value.replace(',', ' ')}
        elif isinstance(field, DateTimeField):
            return {field.code: self._format_date_time(value)}
        else:
            return {field.code: value}

    def _format_field_answers_for(self, fields, submission_values):
        answer_dict = {}
        for field in fields:
            answer = submission_values.get(field.code, '')
            if isinstance(field, SelectField) and field.has_other and isinstance(answer, list) and answer[0] == 'other':
                if field.is_single_select:
                    answer_dict.update({field.code + "_other": answer[1]})
                    answer_dict.update({field.code: answer[0]})
                else:
                    other_selection = []
                    choice_selections = ['other']
                    for item in answer[1].split(' '):
                        if _is_choice_item_in_choice_list(item, field.options):
                            choice_selections.append(item)
                        else:
                            other_selection.append(item)
                    answer_dict.update({field.code + "_other": ' '.join(other_selection)})
                    answer_dict.update({field.code: ' '.join(choice_selections)})
            elif isinstance(field, SelectField)  and field.is_single_select:
                answer_dict.update({field.code: answer if _is_choice_item_in_choice_list(answer, field.options) else ""})
            else:
                answer_dict.update(self.get_dict(field, answer))
        return answer_dict

    def get_model_edit_str(self, form_model_fields, submission_values, project_name, form_code):
        # todo instead of using form_model fields, use xform to find the fields
        answer_dict, edit_model_dict = {'form_code': form_code}, {}
        formatted_field_values = convert_date_values(form_model_fields, self._format_field_answers_for(form_model_fields, submission_values))
        answer_dict.update(formatted_field_values)
        edit_model_dict.update({project_name: answer_dict})
        return re.sub(r'\n', r'\\n', xmldict.dict_to_xml(edit_model_dict))

    def _format_date_time(self, value):
        return datetime.strptime(value, '%d.%m.%Y %H:%M:%S').strftime('%Y-%m-%dT%H:%M:%S')


def convert_date_values(fields, value_dict):
    for field in fields:
        if type(field) is FieldSet:
            dicts = []
            for value in value_dict.get(field.code):
                converted_dict = convert_date_values(field.fields, value)
                dicts.append(converted_dict)
            value_dict[field.code] = dicts
        elif type(field) is DateField:
            converted_value = convert_date_to_ymd(value_dict.get(field.code))
            value_dict[field.code] = converted_value
    return value_dict


def convert_date_to_ymd(str_date):
    if not str_date:
        return str_date
    
    try:
        date = datetime.strptime(str_date, "%Y-%m")
    except ValueError:
        try:
            date = datetime.strptime(str_date, "%Y")
        except ValueError:
            date = datetime.strptime(str_date, "%Y-%m-%d")

    return datetime.strftime(date, "%Y-%m-%d")


class XFormImageProcessor():
    media_fields = [PhotoField, VideoField, AudioField]

    def get_media(self, field, value):
        if type(field) is FieldSet:
            file_names = []
            for v in value:
                for f in field.fields:
                    if isinstance(f, MediaField) and v.get(f.code, ''):
                        file_names.append(v.get(f.code, ''))
            return file_names
        if isinstance(field, MediaField):
            return [value]

    def get_media_files_str(self, form_model_fields, submission_values):
        media_files = []
        for f in form_model_fields:
            values = submission_values.get(f.code, '')
            if values and (isinstance(f, MediaField) or type(f) is FieldSet):
                media_files.extend(self.get_media(f, values))
        return ','.join(media_files)


DIR = os.path.dirname(__file__)


class XFormTransformer():
    def __init__(self, xform_as_string):
        self.xform = xform_as_string
        self.xls_folder = os.path.join(DIR, 'xsl')
        self.HTML_FORM = os.path.join(self.xls_folder, 'openrosa2html5form_php5.xsl')
        self.XML_MODEL = os.path.join(self.xls_folder, 'openrosa2xmlmodel.xsl')

    def transform(self):
        model_xslt = etree.XML(open(self.XML_MODEL, 'r').read())
        transform_model = etree.XSLT(model_xslt)
        model_doc = etree.fromstring(self.xform)
        model_tree = transform_model(model_doc)

        form_xslt = etree.XML(open(self.HTML_FORM, 'r').read())
        transform_form = etree.XSLT(form_xslt)
        form_doc = etree.fromstring(self.xform)
        form_tree = transform_form(form_doc)

        r = model_tree.getroot()
        r.extend(form_tree.getroot())
        return etree.tostring(r)


class TypeNotSupportedException(Exception):
    def __init__(self, message):
        self.message = message

    def __str__(self):
        return self.message


class NestedRepeatsNotSupportedException(Exception):
    def __init__(self):
        self.message = _("nested repeats not supported")

    def __str__(self):
        return self.message


class MultipleLanguagesNotSupportedException(Exception):
    def __init__(self):
        self.message = _("Language specification is not supported")

    def __str__(self):
        return self.message


class LabelForChoiceNotPresentException(Exception):
    def __init__(self, choice_name):
        self.message = _("Label mandatory for choice option with name %s") % choice_name

    def __str__(self):
        return self.message


class PrefetchCSVNotSupportedException(Exception):
    def __init__(self):
        self.message = _("Prefetch of csv not supported")

    def __str__(self):
        return self.message


class LabelForFieldNotPresentException(Exception):
    def __init__(self, field_name):
        self.message = _("Label mandatory for question with name [%s]") % field_name

    def __str__(self):
        return self.message


class ForbiddenWordInFieldNameException(Exception):
    def __init__(self, field_name):
        self.message = _("Field name can't be [%s]") % field_name

    def __str__(self):
        return self.message


class ErrorSyntaxInRelevantException(Exception):
    def __init__(self, field_name, specific):
        if not specific:
            self.message = _("Syntax error in relevant column: [%s]") % field_name
        else:
            self.message = _("Syntax error in relevant column: [%s] near [%s]") % (field_name, specific)

    def __str__(self):
        return self.message


class UniqueIdNotFoundException(Exception):
    def __init__(self, unique_id_type):
        unique_id_type = unique_id_type.title()
        self.message = _("Creating new Identification Number Types using XLSForms. The Identification Number Type %s is "
                         "not yet added but you can add it <a href='/entity/subjects' "
                         "target='_blank'>here</a>.") % unique_id_type


class UniqueIdNotMentionedException(Exception):
    def __init__(self, field_name):
        self.message = _("The Identification Number Type (dw_idnr) is missing in the Constraints column. "
                         "Please add the Identification Number Type and upload again. "
                         "<a target='_blank'>Learn More</a> about how to manage Identification Numbers with XLSForms. ")

    def __str__(self):
        return self.message


class UniqueIdNumbersNotFoundException(Exception):
    def __init__(self, unique_id_type):
        unique_id_type = unique_id_type.title()
        self.message = _(
            "You have not registered a %s yet. Making submission via Smartphone will not be possible. Register a <a href='/entity/subject/create/%s/?web_view=True'>%s</a>") % (
                           unique_id_type, unique_id_type.lower(), unique_id_type)

    def __str__(self):
        return self.message


class XlsProjectParser(XlsParser):
    def parse(self, xls_contents):
        assert xls_contents is not None
        workbook = xlrd.open_workbook(file_contents=xls_contents)
        xls_dict = {}
        for worksheet in workbook.sheets():
            parsedData = []
            for row_num in range(0, worksheet.nrows):
                row = worksheet.row_values(row_num)
                row = self._clean(row)
                parsedData.append(row)
            xls_dict[worksheet.name] = parsedData
        return xls_dict


class XlsParserResponse():
    def __init__(self, errors=None, xform_as_string=None, json_xform_data=None, info=None, is_multiple_languages=False, csv_path=None):
        self.is_multiple_languages = is_multiple_languages
        self.xform_as_string = xform_as_string
        self.json_xform_data = json_xform_data
        if not info:
            info = []
        if not errors:
            errors = []
        self.info = info
        self.errors = errors
        self.csv_path = csv_path

    @property
    def xform_as_string(self):
        return self.xform_as_string if self.xform_as_string else ''

    @property
    def json_xform_data(self):
        return self.json_xform_data

    @property
    def is_multiple_languages(self):
        return self.is_multiple_languages

    @property
    def info(self):
        return self.info
