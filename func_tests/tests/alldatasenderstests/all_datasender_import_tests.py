import os

from django.test import Client
from nose.plugins.attrib import attr
from openpyxl import load_workbook

from framework.base_test import HeadlessRunnerTest, teardown_driver
from pages.globalnavigationpage.global_navigation_page import GlobalNavigationPage
from pages.loginpage.login_page import login
from framework.utils.common_utils import random_number

workbook_row1 = {
    'ds_name': "FName" + random_number(2),
    'ds_number': random_number(6),
    'ds_location': "Loc" + random_number(2),
    'gps_coordinates': '%.1f, ' % float(random_number(1)) + '%.1f' % float(random_number(1)),
    'ds_email': 'DsEmail' + random_number(3) + "@dw.com"
}

workbook_row2 = {
    'ds_name': "FName" + random_number(2),
    'ds_number': random_number(6),
    'ds_location': "Loc" + random_number(2),
    'gps_coordinates': '%.1f, ' % float(random_number(1)) + '%.1f' % float(random_number(1)),
    'ds_email': 'DsEmail' + random_number(3) + "@dw.com"
}

workbook_row3 = {
    'ds_name': "FName" + random_number(2),
    'ds_number': random_number(6),
    'ds_location': "Loc" + random_number(2),
    'gps_coordinates': '%.1f, ' % float(random_number(1)) + '%.1f' % float(random_number(1)),
    'ds_email': 'DsEmail' + random_number(3) + "@dw.com"
}

workbook_row4 = {
    'ds_name': "FName" + random_number(2),
    'ds_number': random_number(6),
    'ds_location': "Loc" + random_number(2),
    'gps_coordinates': '%.1f, ' % float(random_number(1)) + '%.1f' % float(random_number(1)),
    'ds_email': 'DsEmail' + random_number(3) + "@dw.com"
}

class TestAllDataSendersImport(HeadlessRunnerTest):
    @classmethod
    def setUpClass(cls):
        HeadlessRunnerTest.setUpClassFirefox()
        login(cls.driver)

    @classmethod
    def tearDownClass(cls):
        teardown_driver(cls.driver)

    def _upload_imported_datasender(self, file_name, file_path):
        client = Client()
        client.login(username='tester150411@gmail.com', password='tester150411')
        response = client.post(
            path='/entity/datasenders/?qqfile=' + file_name,
            data=open(file_path, 'r').read(),
            content_type='application/octet-stream')
        self.assertEquals(response.status_code, 200)

    def _assert_datasender_uploaded(self, all_datasender_page, workbook):
        all_datasender_page.search_with(workbook['ds_name'])
        self.assertEqual(workbook['ds_name'], all_datasender_page.get_cell_value(1, 2))
        self.assertIn(workbook['ds_location'], all_datasender_page.get_cell_value(1, 5))
        self.assertEqual(workbook['gps_coordinates'], all_datasender_page.get_cell_value(1, 6))
        self.assertEqual(workbook['ds_number'], all_datasender_page.get_cell_value(1, 3))
        self.assertEqual(workbook['ds_email'], all_datasender_page.get_cell_value(1, 4))

    def _prepare_workbook(self):
        file_name = 'DataWinners_ImportContacts.xlsx'
        DIR = os.path.dirname(__file__) + '/'
        file_path = os.path.join(DIR, file_name)
        workbook = load_workbook(file_path)
        worksheet = workbook.get_sheet_by_name('Import Contacts')
        worksheet['A2'] = workbook_row1['ds_name']
        worksheet['B2'] = workbook_row1['ds_number']
        worksheet['C2'] = workbook_row1['ds_location']
        worksheet['D2'] = workbook_row1['gps_coordinates']
        worksheet['E2'] = workbook_row1['ds_email']

        worksheet['A3'] = workbook_row2['ds_name']
        worksheet['B3'] = workbook_row2['ds_number']
        worksheet['C3'] = workbook_row2['ds_location']
        worksheet['D3'] = workbook_row2['gps_coordinates']
        worksheet['E3'] = workbook_row2['ds_email']
        return file_name, file_path, workbook

    def _prepare_update_workbook(self):
        file_name = 'DataWinners_ImportContacts.xlsx'
        DIR = os.path.dirname(__file__) + '/'
        file_path = os.path.join(DIR, file_name)
        workbook = load_workbook(file_path)
        worksheet = workbook.get_sheet_by_name('Import Contacts')
        worksheet['A2'] = workbook_row3['ds_name']
        worksheet['B2'] = workbook_row3['ds_number']
        worksheet['C2'] = workbook_row3['ds_location']
        worksheet['D2'] = workbook_row3['gps_coordinates']
        worksheet['E2'] = workbook_row3['ds_email']

        worksheet['A3'] = workbook_row4['ds_name']
        worksheet['B3'] = workbook_row4['ds_number']
        worksheet['C3'] = workbook_row4['ds_location']
        worksheet['D3'] = workbook_row4['gps_coordinates']
        worksheet['E3'] = workbook_row4['ds_email']

        return file_name, file_path, workbook

    @attr('functional_test')
    def test_should_import_more_than_one_data_sender(self):
        file_name, file_path, workbook = self._prepare_workbook()

        workbook.save(file_path)

        global_navigation = GlobalNavigationPage(self.driver)
        all_datasender_page = global_navigation.navigate_to_all_data_sender_page()

        self._upload_imported_datasender(file_name, file_path)

        self._assert_datasender_uploaded(all_datasender_page, workbook_row1)

        self._assert_datasender_uploaded(all_datasender_page, workbook_row2)

    @attr('functional_test')
    def test_should_import_zand_edit_data_senders(self):
        file_name, file_path, workbook = self._prepare_workbook()
        workbook.save(file_path)
        global_navigation = GlobalNavigationPage(self.driver)
        all_datasender_page = global_navigation.navigate_to_all_data_sender_page()
        self._upload_imported_datasender(file_name, file_path)
        self._assert_datasender_uploaded(all_datasender_page, workbook_row1)
        self._assert_datasender_uploaded(all_datasender_page, workbook_row2)

        file_name, file_path, workbook = self._prepare_update_workbook()
        workbook.save(file_path)
        global_navigation = GlobalNavigationPage(self.driver)
        all_datasender_page = global_navigation.navigate_to_all_data_sender_page()
        self._upload_imported_datasender(file_name, file_path)
        self._assert_datasender_uploaded(all_datasender_page, workbook_row3)
        self._assert_datasender_uploaded(all_datasender_page, workbook_row4)

